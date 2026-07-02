import argparse
import csv
import json
import time
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
import requests
from rdkit import Chem


DEFAULT_INPUT = Path("data/drug.csv")
DEFAULT_OUTPUT = Path("data/drug_ingredient_smiles.csv")
DEFAULT_CACHE = None
PUBCHEM_PROPERTY_URL = (
    "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/"
    "{name}/property/SMILES,CanonicalSMILES,IsomericSMILES/JSON"
)


def split_ingredients(value: Any) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def column_to_index(column: str | int, headers: list[str] | None = None) -> int:
    if isinstance(column, int):
        return column - 1

    value = str(column).strip()
    if not value:
        raise ValueError("ingredient column cannot be empty")

    if value.isdigit():
        return int(value) - 1

    if headers:
        normalized = {header.strip().lower(): index for index, header in enumerate(headers)}
        header_index = normalized.get(value.lower())
        if header_index is not None:
            return header_index

    if value.isalpha():
        index = 0
        for char in value.upper():
            index = index * 26 + ord(char) - ord("A") + 1
        return index - 1

    raise ValueError(f"cannot resolve ingredient column: {column}")


def read_csv_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        rows = list(reader)

    if not rows:
        return [], []
    return [str(cell or "") for cell in rows[0]], rows[1:]


def read_xlsx_rows(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()

    if not rows:
        return [], []
    return [str(cell or "") for cell in rows[0]], rows[1:]


def read_drugs(
    path: Path,
    ingredient_column: str,
    sheet_name: str | None = None,
) -> tuple[list[str], list[tuple[list[Any], list[str]]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        headers, raw_rows = read_csv_rows(path)
    elif suffix == ".xlsx":
        headers, raw_rows = read_xlsx_rows(path, sheet_name)
    else:
        raise ValueError(f"unsupported input file type: {path.suffix}")

    ingredient_index = column_to_index(ingredient_column, headers)
    if ingredient_index < 0:
        raise ValueError(f"invalid ingredient column: {ingredient_column}")

    rows: list[tuple[list[Any], list[str]]] = []
    for line_no, row in enumerate(raw_rows, 2):
        if ingredient_index >= len(row):
            print(f"skip line {line_no}: ingredient column is missing")
            continue

        ingredients = split_ingredients(row[ingredient_index])
        if not ingredients:
            print(f"skip line {line_no}: missing active ingredient")
            continue

        rows.append((row, ingredients))

    return headers, rows


def load_cache(path: Path) -> dict[str, str | None]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    return {str(key): value for key, value in data.items()}


def save_cache(path: Path, cache: dict[str, str | None]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(cache, file, ensure_ascii=False, indent=2, sort_keys=True)


def pick_smiles(payload: dict[str, Any]) -> str | None:
    properties = payload.get("PropertyTable", {}).get("Properties", [])
    if not properties:
        return None
    first = properties[0]
    for key in ("IsomericSMILES", "CanonicalSMILES", "SMILES", "ConnectivitySMILES"):
        value = first.get(key)
        if value:
            return str(value)
    return None


def normalize_smiles(smiles: str | None) -> str | None:
    if not smiles:
        return None
    fragment_mol = Chem.MolFromSmiles(smiles)
    if fragment_mol is None:
        return smiles
    return Chem.MolToSmiles(fragment_mol, canonical=True, isomericSmiles=True)


def fetch_smiles(
    session: requests.Session,
    ingredient: str,
    *,
    timeout: int,
    retries: int,
    delay: float,
) -> str | None:
    url = PUBCHEM_PROPERTY_URL.format(name=quote(ingredient, safe=""))
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, timeout=timeout)
            if response.status_code == 404:
                return None
            response.raise_for_status()
            return normalize_smiles(pick_smiles(response.json()))
        except requests.RequestException as exc:
            if attempt >= retries:
                print(f"failed: {ingredient} ({exc})")
                return None
            time.sleep(delay * attempt)
        except (ValueError, KeyError) as exc:
            print(f"bad response: {ingredient} ({exc})")
            return None
    return None


def write_output(
    path: Path,
    headers: list[str],
    rows: list[tuple[list[Any], list[str]]],
    cache: dict[str, str | None],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    output_headers = headers + ["ActiveIngredient_lookup", "SMILES"]
    output_rows = []
    for row, ingredients in rows:
        padded_row = row + [""] * max(0, len(headers) - len(row))
        for ingredient in ingredients:
            normalized_smiles = normalize_smiles(cache.get(ingredient.upper()))
            output_rows.append(padded_row[: len(headers)] + [ingredient, normalized_smiles or ""])

    if path.suffix.lower() == ".xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ingredient_smiles"
        worksheet.append(output_headers)
        for row in output_rows:
            worksheet.append(row)
        worksheet.freeze_panes = "A2"
        workbook.save(path)
        return

    if path.suffix.lower() != ".csv":
        raise ValueError(f"unsupported output file type: {path.suffix}")

    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(output_headers)
        writer.writerows(output_rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Expand active ingredient values from CSV/XLSX and fetch PubChem SMILES."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--ingredient-column",
        default="2",
        help="Active ingredient column. Accepts header name, Excel letter such as C, or 1-based index such as 3.",
    )
    parser.add_argument("--sheet", default=None, help="XLSX sheet name. Defaults to the active sheet.")
    parser.add_argument(
        "--cache",
        type=Path,
        default=DEFAULT_CACHE,
        help="Cache JSON path. Defaults to pubchem_smiles_cache.json next to the output file.",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Ignore any existing cache and fetch all ingredients from scratch.",
    )
    parser.add_argument("--limit", type=int, default=None, help="Only fetch first N unique ingredients.")
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--delay", type=float, default=0.2, help="Seconds between PubChem requests.")
    args = parser.parse_args()
    if args.cache is None:
        args.cache = args.output.with_name("pubchem_smiles_cache.json")

    headers, rows = read_drugs(args.input, args.ingredient_column, args.sheet)
    unique_ingredients = sorted({ingredient.upper(): ingredient for _, items in rows for ingredient in items}.items())
    if args.limit is not None:
        unique_ingredients = unique_ingredients[: args.limit]

    cache = {} if args.no_cache else load_cache(args.cache)
    session = requests.Session()
    session.headers.update({"User-Agent": "medical-drug-smiles-loader/1.0"})

    total = len(unique_ingredients)
    for index, (cache_key, ingredient) in enumerate(unique_ingredients, 1):
        if cache_key in cache:
            continue
        smile = fetch_smiles(
            session,
            ingredient,
            timeout=args.timeout,
            retries=args.retries,
            delay=args.delay,
        )
        cache[cache_key] = smile
        print(f"[{index}/{total}] {ingredient} -> {smile or ''}")
        save_cache(args.cache, cache)
        time.sleep(args.delay)

    write_output(args.output, headers, rows, cache)
    missing = sum(1 for value in cache.values() if not value)
    expanded = sum(len(items) for _, items in rows)
    print(f"done: rows={expanded}, cache={len(cache)}, missing={missing}, output={args.output}")


if __name__ == "__main__":
    main()
