import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from rdkit import Chem
from rdkit import RDLogger


RDLogger.DisableLog("rdApp.*")


def is_excel_column_letter(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]{1,3}", value.strip()))


def resolve_column(ws, column_arg: str, header_row: int) -> int:
    """
    支持两种写法：
    1. A / B / C 这种 Excel 列字母
    2. SMILES / smiles 这种表头名
    """
    column_arg = column_arg.strip()

    if column_arg.isdigit():
        return int(column_arg)

    if is_excel_column_letter(column_arg):
        return column_index_from_string(column_arg.upper())

    target = column_arg.lower()
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        if str(cell.value).strip().lower() == target:
            return cell.column

    headers = [
        str(cell.value).strip()
        for cell in ws[header_row]
        if cell.value is not None
    ]

    raise ValueError(
        f"找不到列：{column_arg}\n"
        f"当前表头包括：{headers[:30]}"
    )


def canonicalize_smiles(smiles: str):
    if smiles is None:
        return ""

    smiles = str(smiles).strip()
    if not smiles:
        return ""

    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return ""

    return Chem.MolToSmiles(mol, canonical=True)


def main():
    parser = argparse.ArgumentParser(
        description="将 Excel 中指定 SMILES 列标准化，并新增 canonical_smiles 列"
    )

    parser.add_argument(
        "--input",
        required=True,
        help="输入 Excel 文件路径，例如 test.xlsx"
    )

    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 文件路径，不填则自动生成 *_canonical.xlsx"
    )

    parser.add_argument(
        "--smiles-column",
        required=True,
        help="SMILES 所在列，可以是 A/B/C，也可以是表头名，例如 SMILES"
    )

    parser.add_argument(
        "--sheet",
        default=None,
        help="工作表名称，不填则使用第一个工作表"
    )

    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="表头所在行，默认第 1 行"
    )

    parser.add_argument(
        "--new-column-name",
        default="canonical_smiles",
        help="新增标准化列的列名，默认 canonical_smiles"
    )

    args = parser.parse_args()

    input_path = Path(args.input)

    if not input_path.exists():
        print(f"输入文件不存在：{input_path}", file=sys.stderr)
        sys.exit(1)

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.with_name(
            input_path.stem + "_canonical.xlsx"
        )

    wb = load_workbook(input_path)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"找不到工作表：{args.sheet}", file=sys.stderr)
            print(f"可用工作表：{wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        ws = wb[args.sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    smiles_col = resolve_column(ws, args.smiles_column, args.header_row)

    new_col = ws.max_column + 1
    new_col_letter = get_column_letter(new_col)

    ws.cell(row=args.header_row, column=new_col).value = args.new_column_name
    ws.cell(row=args.header_row, column=new_col).font = Font(bold=True)

    total = 0
    success = 0
    failed = 0

    for row in range(args.header_row + 1, ws.max_row + 1):
        raw_smiles = ws.cell(row=row, column=smiles_col).value

        if raw_smiles is None or str(raw_smiles).strip() == "":
            ws.cell(row=row, column=new_col).value = ""
            continue

        total += 1

        canonical = canonicalize_smiles(raw_smiles)

        if canonical:
            success += 1
        else:
            failed += 1

        ws.cell(row=row, column=new_col).value = canonical

    ws.column_dimensions[new_col_letter].width = 45

    wb.save(output_path)

    print("处理完成")
    print(f"输入文件：{input_path}")
    print(f"输出文件：{output_path}")
    print(f"工作表：{ws.title}")
    print(f"SMILES列：{args.smiles_column}")
    print(f"新增列：{args.new_column_name}")
    print(f"总SMILES数：{total}")
    print(f"成功标准化：{success}")
    print(f"失败或非法SMILES：{failed}")


if __name__ == "__main__":
    main()