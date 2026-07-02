# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import hashlib
import sys
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


def add_local_dependency_path() -> None:
    """Allow the repo-local dependency folder used by Codex runs to be reused."""
    repo_root = Path(__file__).resolve().parents[1]
    local_deps = repo_root / ".codex_deps" / "python"
    if local_deps.exists():
        sys.path.insert(0, str(local_deps))


def import_rdkit() -> tuple[Any, Any, Any]:
    try:
        from rdkit import Chem, RDLogger
        from rdkit.Chem import Draw, rdDepictor
    except ImportError:
        add_local_dependency_path()
        try:
            from rdkit import Chem, RDLogger
            from rdkit.Chem import Draw, rdDepictor
        except ImportError as exc:
            raise SystemExit(
                "RDKit is required to draw 2D structures. Install it first, for example:\n"
                "  python -m pip install rdkit\n"
                "or install into the repo-local dependency folder:\n"
                "  python -m pip install --target .codex_deps/python rdkit"
            ) from exc

    RDLogger.DisableLog("rdApp.*")
    return Chem, Draw, rdDepictor


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read an Excel workbook with a fragment_smiles column, add a 2d column, "
            "embed each SMILES 2D structure image into the matching row, and save a new workbook."
        )
    )
    parser.add_argument("input_excel", type=Path, help="Input .xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path. Defaults to '<input_stem>_with_2d.xlsx' beside the input file.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name. Defaults to the active worksheet.",
    )
    parser.add_argument(
        "--smiles-column",
        default="fragment_smiles",
        help="Header name for the SMILES column. Default: fragment_smiles",
    )
    parser.add_argument(
        "--image-column",
        default="2d",
        help="Header name for the output image column. Default: 2d",
    )
    parser.add_argument(
        "--width",
        type=int,
        default=170,
        help="Image width in pixels. Default: 170",
    )
    parser.add_argument(
        "--height",
        type=int,
        default=120,
        help="Image height in pixels. Default: 120",
    )
    parser.add_argument(
        "--keep-images-dir",
        type=Path,
        default=None,
        help="Optional directory to keep generated PNG images. By default temporary images are removed after saving.",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=500,
        help="Print progress every N rows. Use 0 to disable. Default: 500",
    )
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def find_header_column(headers: list[Any], header_name: str) -> int:
    normalized = [normalize_header(value) for value in headers]
    try:
        return normalized.index(header_name) + 1
    except ValueError as exc:
        raise SystemExit(f"Column '{header_name}' was not found. Headers: {normalized}") from exc


def find_optional_header_column(headers: list[Any], header_name: str) -> int | None:
    normalized = [normalize_header(value) for value in headers]
    if header_name not in normalized:
        return None
    return normalized.index(header_name) + 1


def image_anchor_column(image: XLImage) -> int | None:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    col = getattr(marker, "col", None)
    if col is None:
        return None
    return int(col) + 1


def remove_images_in_column(ws: Any, column_index: int) -> int:
    kept = []
    removed = 0
    for image in getattr(ws, "_images", []):
        if image_anchor_column(image) == column_index:
            removed += 1
        else:
            kept.append(image)
    ws._images = kept
    return removed


def prepare_image_column(ws: Any, header_name: str) -> tuple[int, int]:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    normalized = [normalize_header(value) for value in headers]
    if header_name in normalized:
        image_col = normalized.index(header_name) + 1
        removed = remove_images_in_column(ws, image_col)
        for row in range(2, ws.max_row + 1):
            ws.cell(row, image_col).value = None
    else:
        image_col = ws.max_column + 1
        removed = 0

    header = ws.cell(1, image_col)
    header.value = header_name
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="1F4E78")
    header.alignment = Alignment(horizontal="center", vertical="center")
    return image_col, removed


def draw_smiles_png(
    smiles: str,
    png_path: Path,
    width: int,
    height: int,
    Chem: Any,
    Draw: Any,
    rdDepictor: Any,
) -> bool:
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        mol = Chem.MolFromSmarts(smiles)
    if mol is None:
        return False
    rdDepictor.Compute2DCoords(mol)
    image = Draw.MolToImage(mol, size=(width, height), kekulize=True)
    image.save(png_path)
    return True


def iter_column_values(path: Path, sheet_name: str, header_name: str) -> list[Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name]
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        col_index = [normalize_header(value) for value in headers].index(header_name)
    except ValueError as exc:
        raise RuntimeError(f"Column '{header_name}' was not found in {path}") from exc
    return [row[col_index] for row in ws.iter_rows(min_row=2, values_only=True)]


def verify_preserved_column(input_path: Path, output_path: Path, sheet_name: str, header_name: str) -> None:
    input_values = iter_column_values(input_path, sheet_name, header_name)
    output_values = iter_column_values(output_path, sheet_name, header_name)
    if input_values == output_values:
        return

    mismatch_row = None
    for offset, (input_value, output_value) in enumerate(zip(input_values, output_values), start=2):
        if input_value != output_value:
            mismatch_row = (offset, input_value, output_value)
            break
    if mismatch_row is None and len(input_values) != len(output_values):
        mismatch_row = (
            min(len(input_values), len(output_values)) + 2,
            f"{len(input_values)} input values",
            f"{len(output_values)} output values",
        )
    raise RuntimeError(
        f"Column '{header_name}' changed after saving. "
        f"First mismatch: row {mismatch_row[0]}, input={mismatch_row[1]!r}, output={mismatch_row[2]!r}."
    )


def add_2d_images(args: argparse.Namespace) -> Path:
    input_path = args.input_excel.resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file does not exist: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise SystemExit("Only .xlsx files are supported.")

    output_path = (
        args.output.resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_with_2d.xlsx")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    Chem, Draw, rdDepictor = import_rdkit()
    wb = load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    sheet_name = ws.title

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    smiles_col = find_header_column(headers, args.smiles_column)
    drug_number_col = find_optional_header_column(headers, "drug_number")
    original_drug_numbers = (
        [ws.cell(row, drug_number_col).value for row in range(2, ws.max_row + 1)]
        if drug_number_col is not None
        else None
    )
    image_col, removed_images = prepare_image_column(ws, args.image_column)
    image_col_letter = ws.cell(1, image_col).column_letter

    ws.column_dimensions[image_col_letter].width = max(16, round(args.width / 7))
    ws.freeze_panes = ws.freeze_panes or "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, image_col).coordinate}"

    image_cache: dict[str, Path | None] = {}
    invalid: list[tuple[int, str]] = []
    inserted = 0

    temp_context = (
        tempfile.TemporaryDirectory(prefix="fragment_2d_")
        if args.keep_images_dir is None
        else None
    )
    image_dir = args.keep_images_dir or Path(temp_context.name)  # type: ignore[union-attr]
    image_dir.mkdir(parents=True, exist_ok=True)

    try:
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, smiles_col).value
            smiles = str(value).strip() if value is not None else ""
            if not smiles:
                continue

            if smiles not in image_cache:
                digest = hashlib.sha1(smiles.encode("utf-8")).hexdigest()[:16]
                png_path = image_dir / f"{digest}.png"
                if not png_path.exists():
                    ok = draw_smiles_png(
                        smiles,
                        png_path,
                        args.width,
                        args.height,
                        Chem,
                        Draw,
                        rdDepictor,
                    )
                    if not ok:
                        image_cache[smiles] = None
                        invalid.append((row, smiles))
                        continue
                image_cache[smiles] = png_path

            png_path = image_cache[smiles]
            if png_path is None:
                continue

            image = XLImage(str(png_path))
            image.width = args.width
            image.height = args.height
            ws.add_image(image, ws.cell(row, image_col).coordinate)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, args.height * 0.76)
            ws.cell(row, image_col).alignment = Alignment(horizontal="center", vertical="center")
            inserted += 1

            if args.progress_every and inserted % args.progress_every == 0:
                print(f"Inserted {inserted} images...")

        if drug_number_col is not None and original_drug_numbers is not None:
            current_drug_numbers = [ws.cell(row, drug_number_col).value for row in range(2, ws.max_row + 1)]
            if current_drug_numbers != original_drug_numbers:
                raise RuntimeError("drug_number values changed unexpectedly; output was not saved.")

        wb.save(output_path)
        if drug_number_col is not None:
            try:
                verify_preserved_column(input_path, output_path, sheet_name, "drug_number")
            except Exception:
                output_path.unlink(missing_ok=True)
                raise
    finally:
        if temp_context is not None:
            temp_context.cleanup()

    print(f"Input: {input_path}")
    print(f"Output: {output_path}")
    print(f"Worksheet: {ws.title}")
    print(f"Rows: {ws.max_row - 1}")
    print(f"Inserted images: {inserted}")
    print(f"Unique SMILES drawn: {sum(1 for value in image_cache.values() if value is not None)}")
    print(f"Invalid SMILES: {len(invalid)}")
    print(f"Images removed from existing '{args.image_column}' column: {removed_images}")
    if invalid:
        print("Invalid SMILES samples:")
        for row, smiles in invalid[:10]:
            print(f"  row {row}: {smiles}")

    if args.keep_images_dir is not None:
        # Remove unused stale PNG files only when explicitly keeping a directory is requested.
        used = {path.resolve() for path in image_cache.values() if path is not None}
        for file in image_dir.glob("*.png"):
            if file.resolve() not in used:
                file.unlink()

    return output_path


def main() -> None:
    args = parse_args()
    add_2d_images(args)


if __name__ == "__main__":
    main()
