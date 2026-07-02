from __future__ import annotations

from rdkit import Chem
from openpyxl import load_workbook

from bond_enumerated_non_induced_fragments import (
    build_aromatic_systems,
    enumerate_fragments,
    fragment_smiles_from_atom_bond_set,
    parse_smiles,
    write_excel,
)


def fragment_smiles_set(smiles: str, max_atoms: int | None = None) -> set[str]:
    result = enumerate_fragments(smiles, max_atoms=max_atoms)
    return {str(fragment["fragment_smiles"]) for fragment in result["fragments"]}


def test_cyclohexane_non_induced_paths() -> None:
    smiles_set = fragment_smiles_set("C1CCCCC1")
    assert {"CCC", "CCCC", "CCCCC", "CCCCCC", "C1CCCCC1"} <= smiles_set

    mol = parse_smiles("C1CCCCC1")
    assert (
        Chem.MolFragmentToSmiles(
            mol,
            atomsToUse=[0, 1, 2, 3, 4, 5],
            canonical=True,
            isomericSmiles=True,
        )
        == "C1CCCCC1"
    )
    assert fragment_smiles_from_atom_bond_set(mol, range(6), range(5)) == "CCCCCC"


def test_benzene_aromatic_system_is_not_split() -> None:
    smiles_set = fragment_smiles_set("c1ccccc1")
    assert smiles_set == {"c1ccccc1"}
    assert not ({"ccc", "cccc", "ccccc"} & smiles_set)


def test_ethoxybenzene_keeps_aromatic_system_whole() -> None:
    smiles_set = fragment_smiles_set("CCOc1ccccc1")
    assert {"CCO", "c1ccccc1", "Oc1ccccc1", "COc1ccccc1", "CCOc1ccccc1"} <= smiles_set
    assert not any("ccc" in fragment and "c1ccccc1" not in fragment for fragment in smiles_set)


def test_fused_aromatic_system_is_one_protected_unit() -> None:
    mol = parse_smiles("c1ccc2ccccc2c1")
    aromatic_systems = build_aromatic_systems(mol)
    assert len(aromatic_systems) == 1
    assert len(aromatic_systems[0]["atoms"]) == 10

    smiles_set = fragment_smiles_set("c1ccc2ccccc2c1")
    assert smiles_set == {"c1ccc2ccccc2c1"}


def test_deduplicates_without_parent_position() -> None:
    result = enumerate_fragments("CCOCC", min_atoms=4, max_atoms=4)
    smiles = [str(fragment["fragment_smiles"]) for fragment in result["fragments"]]
    assert smiles.count("CCOC") == 1
    assert "COCC" not in smiles


def test_fragment_smiles_are_canonicalized_after_extraction() -> None:
    smiles = (
        "CC(=O)O[C@H]1CC[C@@]2(C)C(=CC[C@@H]3[C@@H]2CC[C@]2(C)"
        "C(c4cccnc4)=CC[C@@H]32)C1"
    )
    result = enumerate_fragments(smiles, max_atoms=5)
    fragment_smiles = [str(fragment["fragment_smiles"]) for fragment in result["fragments"]]
    assert "CCCOC" in fragment_smiles
    assert "COCCC" not in fragment_smiles


def test_excel_output_omits_atom_and_bond_indices(tmp_path) -> None:
    output_path = tmp_path / "fragments.xlsx"
    result = enumerate_fragments("CCOCC")
    write_excel(result, output_path)

    workbook = load_workbook(output_path, read_only=True)
    headers = [
        cell.value
        for cell in next(workbook["fragments"].iter_rows(min_row=1, max_row=1))
    ]
    assert headers == ["index", "fragment_smiles"]
