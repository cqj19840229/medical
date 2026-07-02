from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict, deque
from pathlib import Path
from typing import Iterable

from rdkit import Chem
from rdkit import RDLogger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    Workbook = None
    Alignment = None
    Font = None


RDLogger.DisableLog("rdApp.*")

FRAGMENT_TYPE = "bond_enumerated_non_induced_aromatic_protected"
CHON_ATOMIC_NUMBERS = {1, 6, 7, 8}


def parse_smiles(smiles: str) -> Chem.Mol:
    mol = Chem.MolFromSmiles((smiles or "").strip())
    if mol is None:
        raise ValueError("SMILES cannot be parsed.")
    return mol


def heavy_atom_indices(mol: Chem.Mol) -> list[int]:
    return [atom.GetIdx() for atom in mol.GetAtoms() if atom.GetAtomicNum() > 1]


def is_non_chon_hetero_atom(atom: Chem.Atom) -> bool:
    return atom.GetAtomicNum() not in CHON_ATOMIC_NUMBERS


def non_chon_hetero_atom_symbols(mol: Chem.Mol, atom_indices: Iterable[int]) -> list[str]:
    symbols = {
        mol.GetAtomWithIdx(atom_idx).GetSymbol()
        for atom_idx in atom_indices
        if is_non_chon_hetero_atom(mol.GetAtomWithIdx(atom_idx))
    }
    return sorted(symbols)


def heavy_bond_indices(mol: Chem.Mol) -> list[int]:
    return [
        bond.GetIdx()
        for bond in mol.GetBonds()
        if bond.GetBeginAtom().GetAtomicNum() > 1 and bond.GetEndAtom().GetAtomicNum() > 1
    ]


def build_aromatic_systems(mol: Chem.Mol) -> list[dict[str, set[int]]]:
    aromatic_atoms = {
        atom.GetIdx()
        for atom in mol.GetAtoms()
        if atom.GetIsAromatic() and atom.GetAtomicNum() > 1
    }
    aromatic_bonds = {
        bond.GetIdx()
        for bond in mol.GetBonds()
        if bond.GetIsAromatic()
        and bond.GetBeginAtomIdx() in aromatic_atoms
        and bond.GetEndAtomIdx() in aromatic_atoms
    }

    graph: dict[int, set[int]] = {atom_idx: set() for atom_idx in aromatic_atoms}
    for bond_idx in aromatic_bonds:
        bond = mol.GetBondWithIdx(bond_idx)
        begin_idx = bond.GetBeginAtomIdx()
        end_idx = bond.GetEndAtomIdx()
        graph[begin_idx].add(end_idx)
        graph[end_idx].add(begin_idx)

    systems: list[dict[str, set[int]]] = []
    visited: set[int] = set()
    for atom_idx in sorted(aromatic_atoms):
        if atom_idx in visited:
            continue
        stack = [atom_idx]
        component_atoms: set[int] = set()
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component_atoms.add(current)
            stack.extend(sorted(graph[current] - visited))

        component_bonds = {
            bond_idx
            for bond_idx in aromatic_bonds
            if mol.GetBondWithIdx(bond_idx).GetBeginAtomIdx() in component_atoms
            and mol.GetBondWithIdx(bond_idx).GetEndAtomIdx() in component_atoms
        }
        systems.append({"atoms": component_atoms, "bonds": component_bonds})

    return systems


def atoms_from_bonds(mol: Chem.Mol, bond_indices: Iterable[int]) -> set[int]:
    atom_indices: set[int] = set()
    for bond_idx in bond_indices:
        bond = mol.GetBondWithIdx(int(bond_idx))
        atom_indices.add(bond.GetBeginAtomIdx())
        atom_indices.add(bond.GetEndAtomIdx())
    return atom_indices


def close_protected_aromatic_systems(
    selected_atoms: Iterable[int],
    selected_bonds: Iterable[int],
    aromatic_systems: list[dict[str, set[int]]],
) -> tuple[set[int], set[int], list[int]]:
    closed_atoms = set(selected_atoms)
    closed_bonds = set(selected_bonds)
    protected_units_hit: set[int] = set()

    changed = True
    while changed:
        changed = False
        for unit_index, system in enumerate(aromatic_systems):
            system_atoms = system["atoms"]
            system_bonds = system["bonds"]
            if not (closed_atoms & system_atoms or closed_bonds & system_bonds):
                continue
            before_atoms = len(closed_atoms)
            before_bonds = len(closed_bonds)
            closed_atoms.update(system_atoms)
            closed_bonds.update(system_bonds)
            protected_units_hit.add(unit_index)
            if len(closed_atoms) != before_atoms or len(closed_bonds) != before_bonds:
                changed = True

    return closed_atoms, closed_bonds, sorted(protected_units_hit)


def build_bond_adjacency(mol: Chem.Mol, bond_indices: Iterable[int]) -> dict[int, set[int]]:
    bonds_by_atom: dict[int, set[int]] = defaultdict(set)
    heavy_bonds = set(bond_indices)
    for bond_idx in heavy_bonds:
        bond = mol.GetBondWithIdx(bond_idx)
        bonds_by_atom[bond.GetBeginAtomIdx()].add(bond_idx)
        bonds_by_atom[bond.GetEndAtomIdx()].add(bond_idx)

    adjacency: dict[int, set[int]] = {bond_idx: set() for bond_idx in heavy_bonds}
    for incident_bonds in bonds_by_atom.values():
        for bond_idx in incident_bonds:
            adjacency[bond_idx].update(incident_bonds - {bond_idx})
    return adjacency


def fragment_smiles_from_atom_bond_set(
    mol: Chem.Mol,
    atom_indices: Iterable[int],
    bond_indices: Iterable[int],
) -> str:
    fragment_smiles = Chem.MolFragmentToSmiles(
        mol,
        atomsToUse=sorted(atom_indices),
        bondsToUse=sorted(bond_indices),
        canonical=False,
        isomericSmiles=True,
    )
    return canonicalize_fragment_smiles(fragment_smiles)


def canonicalize_fragment_smiles(fragment_smiles: str) -> str:
    fragment_mol = Chem.MolFromSmiles(fragment_smiles)
    if fragment_mol is None:
        return fragment_smiles
    return Chem.MolToSmiles(fragment_mol, canonical=True, isomericSmiles=True)


def _normalize_state(
    mol: Chem.Mol,
    selected_bonds: Iterable[int],
    aromatic_systems: list[dict[str, set[int]]],
) -> tuple[frozenset[int], frozenset[int], tuple[int, ...]]:
    selected_bond_set = set(selected_bonds)
    selected_atom_set = atoms_from_bonds(mol, selected_bond_set)
    closed_atoms, closed_bonds, protected_units_hit = close_protected_aromatic_systems(
        selected_atom_set,
        selected_bond_set,
        aromatic_systems,
    )
    return frozenset(closed_atoms), frozenset(closed_bonds), tuple(protected_units_hit)


def enumerate_bond_fragments_non_induced(
    mol: Chem.Mol,
    min_atoms: int = 3,
    max_atoms: int | None = None,
    limit: int | None = None,
) -> list[dict[str, object]]:
    if max_atoms is None:
        max_atoms = len(heavy_atom_indices(mol))
    if min_atoms < 1:
        raise ValueError("min_atoms must be >= 1.")
    if max_atoms < min_atoms:
        return []

    candidate_bonds = heavy_bond_indices(mol)
    bond_atom_masks: dict[int, int] = {}
    atom_symbol_by_idx = {
        atom.GetIdx(): atom.GetSymbol()
        for atom in mol.GetAtoms()
        if atom.GetAtomicNum() > 1
    }
    for bond_idx in candidate_bonds:
        bond = mol.GetBondWithIdx(bond_idx)
        bond_atom_masks[bond_idx] = (1 << bond.GetBeginAtomIdx()) | (1 << bond.GetEndAtomIdx())

    bond_adjacency_sets = build_bond_adjacency(mol, candidate_bonds)
    bond_adjacency_masks = {
        bond_idx: sum(1 << neighbor_idx for neighbor_idx in neighbors)
        for bond_idx, neighbors in bond_adjacency_sets.items()
    }

    aromatic_systems = build_aromatic_systems(mol)
    aromatic_system_masks = [
        (
            sum(1 << atom_idx for atom_idx in system["atoms"]),
            sum(1 << bond_idx for bond_idx in system["bonds"]),
        )
        for system in aromatic_systems
    ]
    atom_mask_cache: dict[int, int] = {}
    closure_cache: dict[int, tuple[int, int, tuple[int, ...]]] = {}
    visited_states: set[int] = set()
    queued_states: set[int] = set()
    queue: deque[int] = deque()
    unique: dict[str, dict[str, object]] = {}

    def iter_mask_indices(mask: int) -> Iterable[int]:
        while mask:
            bit = mask & -mask
            yield bit.bit_length() - 1
            mask ^= bit

    def atom_mask_from_bond_mask(bond_mask: int) -> int:
        cached = atom_mask_cache.get(bond_mask)
        if cached is not None:
            return cached
        atom_mask = 0
        for bond_idx in iter_mask_indices(bond_mask):
            atom_mask |= bond_atom_masks[bond_idx]
        atom_mask_cache[bond_mask] = atom_mask
        return atom_mask

    def close_bond_mask(bond_mask: int) -> tuple[int, int, tuple[int, ...]]:
        cached = closure_cache.get(bond_mask)
        if cached is not None:
            return cached

        atom_mask = atom_mask_from_bond_mask(bond_mask)
        closed_bond_mask = bond_mask
        protected_units_hit: set[int] = set()
        changed = True
        while changed:
            changed = False
            for unit_index, (system_atom_mask, system_bond_mask) in enumerate(aromatic_system_masks):
                if not (atom_mask & system_atom_mask or closed_bond_mask & system_bond_mask):
                    continue
                before_atom_mask = atom_mask
                before_bond_mask = closed_bond_mask
                atom_mask |= system_atom_mask
                closed_bond_mask |= system_bond_mask
                protected_units_hit.add(unit_index)
                if atom_mask != before_atom_mask or closed_bond_mask != before_bond_mask:
                    changed = True

        closed = (atom_mask, closed_bond_mask, tuple(sorted(protected_units_hit)))
        closure_cache[bond_mask] = closed
        return closed

    def atom_indices_from_mask(atom_mask: int) -> list[int]:
        return list(iter_mask_indices(atom_mask))

    def bond_indices_from_mask(bond_mask: int) -> list[int]:
        return list(iter_mask_indices(bond_mask))

    def non_chon_symbols_from_atom_mask(atom_mask: int) -> list[str]:
        symbols = {
            atom_symbol_by_idx[atom_idx]
            for atom_idx in iter_mask_indices(atom_mask)
            if is_non_chon_hetero_atom(mol.GetAtomWithIdx(atom_idx))
        }
        return sorted(symbols)

    def enqueue(bond_mask: int) -> None:
        closed_atom_mask, closed_bond_mask, _protected_units_hit = close_bond_mask(bond_mask)
        if closed_atom_mask.bit_count() > max_atoms:
            return
        if closed_bond_mask not in visited_states and closed_bond_mask not in queued_states:
            queued_states.add(closed_bond_mask)
            queue.append(closed_bond_mask)

    for seed_bond_idx in sorted(candidate_bonds):
        enqueue(1 << seed_bond_idx)

    while queue:
        state = queue.popleft()
        queued_states.discard(state)
        if state in visited_states:
            continue
        visited_states.add(state)

        selected_atom_mask, selected_bond_mask, protected_units_hit = close_bond_mask(state)
        atom_count = selected_atom_mask.bit_count()
        if min_atoms <= atom_count <= max_atoms:
            selected_atoms = atom_indices_from_mask(selected_atom_mask)
            selected_bonds = bond_indices_from_mask(selected_bond_mask)
            fragment_smiles = fragment_smiles_from_atom_bond_set(
                mol,
                selected_atoms,
                selected_bonds,
            )
            selected_atoms = sorted(selected_atoms)
            selected_bonds = sorted(selected_bonds)
            existing = unique.get(fragment_smiles)
            if existing is None or (
                selected_atoms,
                selected_bonds,
            ) < (
                existing["_atom_indices"],
                existing["_bond_indices"],
            ):
                record = {
                    "fragment_smiles": fragment_smiles,
                    "_atom_indices": selected_atoms,
                    "_bond_indices": selected_bonds,
                    "fragment_type": FRAGMENT_TYPE,
                    "protected_units_hit": list(protected_units_hit),
                    "protected_units_hit_json": json.dumps(list(protected_units_hit), ensure_ascii=False),
                    "atom_count": atom_count,
                    "bond_count": len(selected_bonds),
                    "heavy_atoms": atom_count,
                    "non_chon_hetero_atoms": non_chon_symbols_from_atom_mask(selected_atom_mask),
                }
                unique[fragment_smiles] = record
                if limit is not None and len(unique) >= limit:
                    break

        frontier_mask = 0
        for bond_idx in iter_mask_indices(selected_bond_mask):
            frontier_mask |= bond_adjacency_masks.get(bond_idx, 0)
        for next_bond_idx in iter_mask_indices(frontier_mask & ~selected_bond_mask):
            if atom_count >= max_atoms and bond_atom_masks[next_bond_idx] & ~selected_atom_mask:
                continue
            enqueue(selected_bond_mask | (1 << next_bond_idx))

    fragments = sorted(
        unique.values(),
        key=lambda item: (
            int(item["atom_count"]),
            str(item["fragment_smiles"]),
            str(item["_atom_indices"]),
            str(item["_bond_indices"]),
        ),
    )
    return fragments[:limit] if limit is not None else fragments


def enumerate_fragments(
    smiles: str,
    limit: int | None = None,
    min_atoms: int = 3,
    max_atoms: int | None = None,
) -> dict[str, object]:
    mol = parse_smiles(smiles)
    canonical_smiles = Chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    heavy_atoms = heavy_atom_indices(mol)
    effective_max_atoms = len(heavy_atoms) if max_atoms is None else max_atoms
    fragments = enumerate_bond_fragments_non_induced(
        mol,
        min_atoms=min_atoms,
        max_atoms=effective_max_atoms,
        limit=limit,
    )
    return {
        "input_smiles": smiles,
        "canonical_smiles": canonical_smiles,
        "heavy_atom_count": len(heavy_atoms),
        "min_atoms": min_atoms,
        "max_atoms": effective_max_atoms,
        "fragment_type": FRAGMENT_TYPE,
        "non_chon_hetero_atoms": non_chon_hetero_atom_symbols(mol, heavy_atoms),
        "fragments": fragments,
    }


def public_fragment_record(fragment: dict[str, object]) -> dict[str, object]:
    return {
        "fragment_smiles": fragment["fragment_smiles"],
        "atom_count": fragment["atom_count"],
        "bond_count": fragment["bond_count"],
        "fragment_type": fragment["fragment_type"],
        "protected_units_hit": fragment["protected_units_hit"],
        "protected_units_hit_json": fragment["protected_units_hit_json"],
        "heavy_atoms": fragment["heavy_atoms"],
        "non_chon_hetero_atoms": fragment["non_chon_hetero_atoms"],
    }


def write_json(result: dict[str, object], output_path: Path) -> None:
    serializable_result = dict(result)
    serializable_result["fragments"] = [
        public_fragment_record(fragment)
        for fragment in result["fragments"]
    ]
    output_path.write_text(
        json.dumps(serializable_result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_csv(result: dict[str, object], output_path: Path) -> None:
    fragments = result["fragments"]
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "index",
                "fragment_smiles",
            ],
        )
        writer.writeheader()
        for index, fragment in enumerate(fragments, start=1):
            writer.writerow(
                {
                    "index": index,
                    "fragment_smiles": fragment["fragment_smiles"],
                }
            )


def write_excel(result: dict[str, object], output_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to write Excel files.")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_rows = [
        ("input_smiles", result["input_smiles"]),
        ("canonical_smiles", result["canonical_smiles"]),
        ("heavy_atom_count", result["heavy_atom_count"]),
        ("fragment_atom_range", f"{result['min_atoms']}..{result['max_atoms']}"),
        ("fragment_type", result["fragment_type"]),
        ("unique_fragments", len(result["fragments"])),
    ]
    for row_idx, (key, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_idx, column=1, value=key)
        summary_sheet.cell(row=row_idx, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 90

    fragment_sheet = workbook.create_sheet("fragments")
    headers = [
        "index",
        "fragment_smiles",
    ]
    fragment_sheet.append(headers)
    for cell in fragment_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, fragment in enumerate(result["fragments"], start=1):
        fragment_sheet.append(
            [
                index,
                fragment["fragment_smiles"],
            ]
        )

    fragment_sheet.freeze_panes = "A2"
    fragment_sheet.column_dimensions["A"].width = 10
    fragment_sheet.column_dimensions["B"].width = 45
    for row in fragment_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(output_path)


def parse_optional_positive_int(value: str) -> int | None:
    parsed = int(value)
    if parsed < 0:
        raise argparse.ArgumentTypeError("value must be >= 0")
    return None if parsed == 0 else parsed


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Enumerate non-induced heavy-atom SMILES fragments by growing bond sets."
    )
    parser.add_argument("smiles", nargs="?", help="Input drug SMILES.")
    parser.add_argument(
        "-o",
        "--output",
        default="bond_enumerated_fragments.xlsx",
        help="Output path. Defaults to .xlsx. Use .json for JSON or .csv for CSV.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional maximum number of unique fragments to write after sorting.",
    )
    parser.add_argument("--min-atoms", type=int, default=3, help="Minimum heavy atoms per fragment.")
    parser.add_argument(
        "--max-atoms",
        type=parse_optional_positive_int,
        default=None,
        help=(
            "Maximum heavy atoms per fragment. Defaults to the input SMILES non-hydrogen "
            "heavy-atom count; use 0 for the same default."
        ),
    )
    args = parser.parse_args()

    if not args.smiles:
        parser.error("smiles is required.")

    result = enumerate_fragments(
        args.smiles,
        limit=args.limit,
        min_atoms=args.min_atoms,
        max_atoms=args.max_atoms,
    )
    output_path = Path(args.output)
    if output_path.suffix.lower() == ".json":
        write_json(result, output_path)
    elif output_path.suffix.lower() == ".csv":
        write_csv(result, output_path)
    else:
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        write_excel(result, output_path)

    print(f"canonical_smiles: {result['canonical_smiles']}")
    print(f"heavy_atom_count: {result['heavy_atom_count']}")
    print(f"fragment_atom_range: {result['min_atoms']}..{result['max_atoms']}")
    print(f"fragment_type: {FRAGMENT_TYPE}")
    print(f"unique_fragments: {len(result['fragments'])}")
    print(f"output: {output_path.resolve()}")


if __name__ == "__main__":
    main()
