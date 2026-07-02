"""
药物分子 SMILES 片段拆解工具
============================
支持两种运行模式：
  1. 批量模式  — 读取 Excel 文件 B 列的 SMILES，逐个拆解，每个药物一个 Sheet
  2. 单分子模式 — 交互式输入单个 SMILES

依赖：pip install rdkit openpyxl
用法：python drug_fragment_decompose.py
"""

from __future__ import annotations
import re
import sys
import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple

import rdkit.RDLogger as RDLogger
RDLogger.DisableLog("rdApp.*")

try:
    from rdkit import Chem
    from rdkit.Chem.rdMolDescriptors import CalcNumRings, CalcNumAromaticRings
except ImportError:
    sys.exit("请先安装 RDKit：pip install rdkit")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("请先安装 openpyxl：pip install openpyxl")


# ─────────────────────────────────────────────
# 样式常量
# ─────────────────────────────────────────────
C_HDR_BG  = "2563EB"
C_HDR_FG  = "FFFFFF"
C_BENZ_BG = "FEF3C7"
C_BENZ_FG = "92400E"
C_ALT_BG  = "F8FAFC"
C_STAT_BG = "EFF6FF"
C_BORDER  = "CBD5E1"
C_TITLE   = "1E3A5F"
C_ERR_BG  = "FEE2E2"
C_ERR_FG  = "991B1B"

BENZENE_PAT = Chem.MolFromSmarts("c1ccccc1")

COLUMNS = [
    ("序号",      7),
    ("SMILES",   46),
    ("重原子数",  10),
    ("环数",       8),
    ("芳香环数",  10),
    ("含苯环",     9),
]


def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────
# 分子解析与片段属性
# ─────────────────────────────────────────────

def parse_mol(smiles: str) -> Tuple[Optional[Chem.Mol], Optional[str]]:
    """盐型自动取重原子最多的组分。"""
    best = None
    for p in smiles.strip().split("."):
        m = Chem.MolFromSmiles(p)
        if m and (best is None or m.GetNumHeavyAtoms() > best[0].GetNumHeavyAtoms()):
            best = (m, p)
    return best if best else (None, None)


def frag_info(smi: str) -> Optional[Dict]:
    """规范化 SMILES 并计算片段属性，不合法片段返回 None。"""
    m = Chem.MolFromSmiles(smi)
    if m is None:
        return None
    return {
        "smiles":         Chem.MolToSmiles(m),
        "heavy_atoms":    m.GetNumHeavyAtoms(),
        "rings":          CalcNumRings(m),
        "aromatic_rings": CalcNumAromaticRings(m),
        "has_benzene":    m.HasSubstructMatch(BENZENE_PAT),
    }


# ─────────────────────────────────────────────
# 滑动窗口（BFS 连通子图枚举）
# ─────────────────────────────────────────────

def run_sliding(mol: Chem.Mol, min_a: int, max_a: int) -> List[Dict]:
    """
    枚举分子中所有大小在 [min_a, max_a] 的连通原子子集。
    - 用规范化 SMILES 去重（同一化学结构只保留一个）
    - 自动过滤 RDKit 无法解析的截断片段（如不完整芳环）
    - max_a 超过分子实际原子数时自动截断
    """
    n = mol.GetNumAtoms()
    real_max = min(max_a, n)

    def bfs_grow(start: int, size: int):
        queue = [frozenset([start])]
        found, visited = set(), set()
        while queue:
            cur = queue.pop()
            if len(cur) == size:
                found.add(cur)
                continue
            for idx in cur:
                for nb in mol.GetAtomWithIdx(idx).GetNeighbors():
                    nxt = frozenset(cur | {nb.GetIdx()})
                    if nxt not in visited:
                        visited.add(nxt)
                        queue.append(nxt)
        return found

    seen_sets = set()
    results   = {}

    for start in range(n):
        for size in range(min_a, real_max + 1):
            for atom_set in bfs_grow(start, size):
                if atom_set in seen_sets:
                    continue
                seen_sets.add(atom_set)
                try:
                    raw = Chem.MolFragmentToSmiles(mol, sorted(atom_set))
                except Exception:
                    continue
                if not raw:
                    continue
                info = frag_info(raw)
                if info is None:
                    continue
                if not (min_a <= info["heavy_atoms"] <= max_a):
                    continue
                canon = info["smiles"]
                if canon not in results:
                    results[canon] = info

    return sorted(results.values(), key=lambda x: (-x["heavy_atoms"], x["smiles"]))


# ─────────────────────────────────────────────
# 从 Excel 读取 SMILES
# ─────────────────────────────────────────────

def read_smiles_from_excel(path: str) -> List[Tuple[str, str]]:
    """
    读取 Excel B 列的 SMILES，A 列作为药物名称。
    自动跳过空行，第一行若不像 SMILES 则视为表头跳过。
    返回 [(name, smiles), ...]
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or len(row) < 2:
            continue
        name_raw   = row[0]
        smiles_raw = row[1]
        if smiles_raw is None:
            continue
        smiles = str(smiles_raw).strip()
        if not smiles:
            continue
        # 跳过表头行：B 列内容不含任何 SMILES 特征字符
        if i == 0 and not any(c in smiles for c in "CNOSPFIcnops()=#@[]"):
            continue
        name = str(name_raw).strip() if name_raw else f"drug_{i + 1}"
        records.append((name, smiles))
    wb.close()
    return records


# ─────────────────────────────────────────────
# Excel 写出工具
# ─────────────────────────────────────────────

def _write_drug_sheet(ws, frags: List[Dict], drug_name: str,
                      input_smi: str, mol_ha: int, min_a: int, max_a: int):
    """向已存在的 worksheet 写入单个药物的片段列表。"""
    # 标题
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value      = f"{drug_name}  ·  滑动窗口  ·  原子数 {min_a}~{max_a}"
    c.font       = Font(name="Arial", size=12, bold=True, color=C_TITLE)
    c.fill       = PatternFill("solid", fgColor=C_STAT_BG)
    c.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # 元信息
    benz_n = sum(1 for f in frags if f.get("has_benzene"))
    meta = [
        ("SMILES",     input_smi[:100] + ("…" if len(input_smi) > 100 else "")),
        ("重原子数",   mol_ha),
        ("原子数范围", f"{min_a} ~ {max_a}"),
        ("片段总数",   len(frags)),
        ("含苯环片段", f"{benz_n} 个（{benz_n/len(frags)*100:.1f}%）" if frags else 0),
    ]
    for i, (k, v) in enumerate(meta):
        r = i + 2
        ws.cell(r, 1, k).font = Font(name="Arial", size=9, bold=True, color="64748B")
        ws.cell(r, 2, v).font = Font(name="Arial", size=9, color="1E293B")
        ws.merge_cells(f"B{r}:F{r}")

    # 表头
    hdr = len(meta) + 3
    hf  = PatternFill("solid", fgColor=C_HDR_BG)
    hfn = Font(name="Arial", size=10, bold=True, color=C_HDR_FG)
    ha  = Alignment(horizontal="center", vertical="center")
    for ci, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(hdr, ci, name)
        c.font = hfn; c.fill = hf; c.alignment = ha; c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hdr].height = 20
    ws.freeze_panes = ws.cell(hdr + 1, 1)
    if frags:
        ws.auto_filter.ref = f"A{hdr}:F{hdr + len(frags)}"

    # 数据行
    benz_fill  = PatternFill("solid", fgColor=C_BENZ_BG)
    alt_fill   = PatternFill("solid", fgColor=C_ALT_BG)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for idx, f in enumerate(frags):
        row = hdr + 1 + idx
        is_benz = f.get("has_benzene", False)
        fill = benz_fill if is_benz else (alt_fill if idx % 2 else white_fill)
        txc  = C_BENZ_FG if is_benz else "1E293B"
        vals   = [idx + 1, f.get("smiles", ""), f.get("heavy_atoms", ""),
                  f.get("rings", ""), f.get("aromatic_rings", ""),
                  "✓" if is_benz else "·"]
        aligns = ["center", "left", "center", "center", "center", "center"]
        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row, ci, val)
            c.font      = Font(name="Arial", size=9, color=txc)
            c.fill      = fill
            c.alignment = Alignment(horizontal=aln, vertical="center",
                                    wrap_text=(ci == 2))
            c.border    = thin_border()
        ws.row_dimensions[row].height = 15


def _write_summary_sheet(ws, summary_rows: List[Dict], min_a: int, max_a: int):
    """写汇总 Sheet，每行一个药物。"""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"批量拆解汇总  ·  滑动窗口  ·  原子数 {min_a}~{max_a}"
    c.font      = Font(name="Arial", size=13, bold=True, color=C_TITLE)
    c.fill      = PatternFill("solid", fgColor=C_STAT_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    HDR = [("序号",6),("药物名称",20),("输入SMILES",46),("重原子数",10),
           ("片段总数",10),("含苯环",10),("含芳香环",10),("状态",10)]
    hf  = PatternFill("solid", fgColor=C_HDR_BG)
    hfn = Font(name="Arial", size=10, bold=True, color=C_HDR_FG)
    ha  = Alignment(horizontal="center", vertical="center")
    for ci, (name, width) in enumerate(HDR, 1):
        c = ws.cell(2, ci, name)
        c.font = hfn; c.fill = hf; c.alignment = ha; c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = ws.cell(3, 1)
    ws.auto_filter.ref = f"A2:H{2 + len(summary_rows)}"

    alt = PatternFill("solid", fgColor=C_ALT_BG)
    wht = PatternFill("solid", fgColor="FFFFFF")
    err = PatternFill("solid", fgColor=C_ERR_BG)

    for i, r in enumerate(summary_rows):
        row    = i + 3
        is_err = r.get("status") != "OK"
        fill   = err if is_err else (alt if i % 2 else wht)
        txc    = C_ERR_FG if is_err else "1E293B"
        smi_s  = r["smiles"][:80] + ("…" if len(r["smiles"]) > 80 else "")
        vals   = [i + 1, r["name"], smi_s, r.get("mol_ha", "—"),
                  r.get("frag_n", 0), r.get("benz_n", 0),
                  r.get("arom_n", 0), r.get("status", "—")]
        aligns = ["center","left","left","center","center","center","center","center"]
        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row, ci, val)
            c.font      = Font(name="Arial", size=9, color=txc)
            c.fill      = fill
            c.alignment = Alignment(horizontal=aln, vertical="center",
                                    wrap_text=(ci == 3))
            c.border    = thin_border()
        ws.row_dimensions[row].height = 15


def _safe_sheet_name(name: str, idx: int, existing: List[str]) -> str:
    """生成合法唯一的 Sheet 名（Excel 限 31 字符，禁用特殊符号）。"""
    clean = re.sub(r'[\[\]:*?/\\]', "_", name)[:28].strip()
    base  = clean or f"drug_{idx}"
    candidate, n = base, 1
    while candidate in existing:
        candidate = f"{base[:25]}_{n}"
        n += 1
    return candidate


# ─────────────────────────────────────────────
# 批量模式
# ─────────────────────────────────────────────

def run_batch(input_xlsx: str, min_a: int, max_a: int, output_xlsx: str):
    print(f"\n{'='*58}")
    print(f"  批量拆解模式")
    print(f"  输入：{input_xlsx}")
    print(f"  原子数范围：{min_a} ~ {max_a}")
    print(f"{'='*58}")

    records = read_smiles_from_excel(input_xlsx)
    if not records:
        print("❌ 未在 B 列找到有效 SMILES，请检查文件。")
        return
    print(f"\n读取到 {len(records)} 条记录\n")

    wb = Workbook()
    wb.remove(wb.active)   # 删除默认空 Sheet

    summary_rows  = []
    existing_names = []

    for idx, (name, smiles) in enumerate(records):
        print(f"  [{idx+1:>3}/{len(records)}] {name:<25}", end="  ", flush=True)

        mol, _ = parse_mol(smiles)
        if mol is None:
            print("❌ 解析失败")
            summary_rows.append({
                "name": name, "smiles": smiles,
                "status": "解析失败", "frag_n": 0, "benz_n": 0, "arom_n": 0,
            })
            safe = _safe_sheet_name(name, idx + 1, existing_names)
            existing_names.append(safe)
            ws_e = wb.create_sheet(safe)
            ws_e["A1"].value = f"SMILES 解析失败：{smiles}"
            ws_e["A1"].font  = Font(name="Arial", size=10, color=C_ERR_FG)
            continue

        mol_ha = mol.GetNumHeavyAtoms()
        frags  = run_sliding(mol, min_a, max_a)
        benz_n = sum(1 for f in frags if f.get("has_benzene"))
        arom_n = sum(1 for f in frags if f.get("aromatic_rings", 0) > 0)
        print(f"重原子={mol_ha:>3}  片段={len(frags):>5}  苯环={benz_n:>4}")

        safe = _safe_sheet_name(name, idx + 1, existing_names)
        existing_names.append(safe)
        ws = wb.create_sheet(safe)
        _write_drug_sheet(ws, frags, name, smiles, mol_ha, min_a, max_a)

        summary_rows.append({
            "name": name, "smiles": smiles, "mol_ha": mol_ha,
            "frag_n": len(frags), "benz_n": benz_n, "arom_n": arom_n,
            "status": "OK",
        })

    # 汇总 Sheet 插到最前面
    ws_sum = wb.create_sheet("汇总", 0)
    _write_summary_sheet(ws_sum, summary_rows, min_a, max_a)

    wb.save(output_xlsx)
    ok_n  = sum(1 for r in summary_rows if r["status"] == "OK")
    err_n = len(summary_rows) - ok_n
    total = sum(r.get("frag_n", 0) for r in summary_rows)
    print(f"\n{'='*58}")
    print(f"✅ 已保存：{output_xlsx}")
    print(f"   成功 {ok_n} 个  失败 {err_n} 个  总片段数 {total}")
    print(f"{'='*58}")


# ─────────────────────────────────────────────
# 单分子模式
# ─────────────────────────────────────────────

def run_single(min_a: int, max_a: int):
    default_smi = (
        "CCCCCCCCCCCC(=O)N[C@@H](CO[C@@H]1O[C@H](CO)"
        "[C@@H](O)[C@H](O)[C@H]1O)[C@@H](O)c1ccc(F)cc1"
        ".OC(CC(O)=O)(CC(O)=O)C(O)=O"
    )
    print(f"\n直接回车使用示例分子（Eliglustat Tartrate）")
    smiles = input("输入 SMILES：").strip() or default_smi

    mol, main_smi = parse_mol(smiles)
    if mol is None:
        print("❌ SMILES 解析失败，请检查输入。")
        return

    print(f"\n主药组分：{main_smi[:65]}{'…' if len(main_smi)>65 else ''}")
    print(f"重原子数：{mol.GetNumHeavyAtoms()}")
    print(f"\n⏳ 正在拆解，原子数范围 {min_a}~{max_a}，稍等...")

    frags = run_sliding(mol, min_a, max_a)
    if not frags:
        print("⚠️  未找到符合条件的片段。")
        return

    print(f"✓ 找到 {len(frags)} 个片段")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_out = f"fragments_{ts}.xlsx"
    out_raw  = input(f"\n输出文件名（默认 {default_out}）：").strip()
    out_path = (out_raw or default_out)
    if not out_path.endswith(".xlsx"):
        out_path += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "片段列表"
    _write_drug_sheet(ws, frags, "单分子", smiles,
                      mol.GetNumHeavyAtoms(), min_a, max_a)
    wb.save(out_path)
    benz_n = sum(1 for f in frags if f.get("has_benzene"))
    print(f"\n✅ 已保存：{out_path}")
    print(f"   片段总数：{len(frags)}  含苯环：{benz_n}")


# ─────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────

def main():
    print("=" * 58)
    print("  药物分子 SMILES 片段拆解工具  （滑动窗口）")
    print("=" * 58)

    print("\n选择运行模式：")
    print("  1  批量模式  — 从 Excel B 列读取 SMILES，逐个拆解")
    print("  2  单分子    — 手动输入一个 SMILES")
    mode = input("\n请输入编号 [1/2]（默认 1）：").strip() or "1"

    print(f"\n设定片段原子数范围（默认 3~22）")
    try:
        min_a = int(input("最小原子数（默认 3）：").strip() or 3)
        max_a = int(input("最大原子数（默认 22）：").strip() or 22)
    except ValueError:
        print("输入无效，使用默认 3~22。")
        min_a, max_a = 3, 22
    if min_a > max_a:
        min_a, max_a = max_a, min_a

    if mode == "2":
        run_single(min_a, max_a)
        return

    # 批量模式
    input_path = input("\n输入 Excel 文件路径：").strip()
    if not input_path:
        print("❌ 未输入文件路径。"); return
    if not os.path.exists(input_path):
        print(f"❌ 文件不存在：{input_path}"); return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_out = f"fragments_batch_{ts}.xlsx"
    out_raw  = input(f"输出文件名（默认 {default_out}）：").strip()
    out_path = (out_raw or default_out)
    if not out_path.endswith(".xlsx"):
        out_path += ".xlsx"

    run_batch(input_path, min_a, max_a, out_path)


if __name__ == "__main__":
    main()
